import geopandas as gpd
import pandas as pd
import numpy as np
from shapely.geometry import Point, LineString
import networkx as nx
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import sys
import os

def get_closest_point(line, point):
    '''
    Calculate the closest point on a line to a given point.

    Parameters
    ----------
    line : shapely.geometry.LineString
        The line on which to find the closest point.
    point : shapely.geometry.Point
        The point from which to find the closest point on the line.

    Returns
    -------
    shapely.geometry.Point
        The closest point on the line to the given point.
    '''
    closest_point = line.interpolate(line.project(point))
    return closest_point

def calculate_GLF(n):
    '''
    Calculate the simultaneity factor (Gleichzeitigkeitsfaktor).

    Parameters
    ----------
    n : int
        Number of buildings.

    Returns
    -------
    float
        The simultaneity factor.
    '''
    a = 0.4497
    b = 0.5512
    c = 53.8483
    d = 1.7627
    return a + (b / (1 + pow(n/c, d)))

def calculate_volumeflow(kW_GLF, htemp, ltemp):
    '''
    Calculate the volumetric flow rate in a pipeline.

    Parameters
    ----------
    kW_GLF : float
        Thermal power with simultaneity factor applied.
    htemp : float
        Supply temperature.
    ltemp : float
        Return temperature.

    Returns
    -------
    float
        Volumetric flow rate in liters per second.
    '''
    #piecewise linear interpolation
    t = [0, 10, 20, 30, 40, 50, 60, 70, 80 , 90, 100] 
    d = [0.99984, 0.9997, 0.99821, 0.99565, 0.99222, 0.98803, 0.9832, 0.97778, 0.97182, 0.96535, 0.9584]
    c = [4.2176, 4.1921, 4.1818, 4.1784, 4.1785, 4.1806, 4.1843, 4.1895, 4.1963, 4.205, 4.2159]
    density = np.interp(int(htemp), t, d)
    cp = np.interp(int(htemp), t, c)

    volumeflow = kW_GLF / (density * cp * (int(htemp) - int(ltemp))) # liter/s
    return volumeflow

def calculate_diameter_velocity_loss(volumeflow, htemp, ltemp, length, pipe_info, edge_type):
    '''
    Calculate the diameter, velocity, and loss of pipelines.

    Parameters
    ----------
    volumeflow : float
        Volumetric flow rate.
    htemp : float
        Supply temperature.
    ltemp : float
        Return temperature.
    length : float
        Length of the pipeline.
    pipe_info : DataFrame
        DataFrame containing pipeline information with columns 'DN', 'di', 'U-Value', 'v_max'.
    edge_type : string
        String containing the type of the edge e.g. 'Hausanschluss'.
        
    Returns
    -------
    tuple
        A tuple containing:
        - DN (float): Nominal diameter.
        - velocity (float): Velocity in the pipeline.
        - loss (float): Heat loss.
        - loss_extra (float): Heat loss with extra insulation.
    '''
    # non-house connections should have at least dn = 32(=DN[2])
    if edge_type == 'Hausanschluss':
        start_index = 0
    else:
        start_index = 2

    # search index of suitable max. Volume Flow
    idx = pipe_info['max_volumeFlow'][start_index:].searchsorted(volumeflow, side='right') + start_index

    # if volumeflow is too high take last DN, which should not happen with DN = 300
    if idx >= len(pipe_info):
        idx = len(pipe_info) - 1

    # get values from pipe_info
    d_i = pipe_info['di'].iloc[idx]
    DN = pipe_info['DN'].iloc[idx]
    u = pipe_info['U-Value'].iloc[idx]
    u_extra = pipe_info['U-Value_extra_insulation'].iloc[idx]

    # calculate velocity and loss
    r = d_i / 2
    velocity = volumeflow * 1000 / (np.pi * pow(r, 2))  # dm^3/mm^2 --> Factor 1000

    # temperature difference for heat loss
    mtemp = (htemp+ltemp)/2
    K = mtemp - 10  # Outside Temp. = 10°C for underground installation 
    
    loss = 8760 * 2 * (u * K * length) / 1000  # 8760 h/a, 2* --> supply and return
    loss_extra = 8760 * 2 * (u_extra * K * length) / 1000
    return DN, velocity, loss, loss_extra

class Streets:
    '''
    A class to manage street geometries and to add connection points from buildings and energy sources to the streets.

    Attributes
    ----------
    gdf : GeoDataFrame
        A GeoDataFrame containing street geometries and attributes.

    Methods
    -------
    add_connection_to_streets(buildings, sources):
        Inserts connection points into the street lines based on buildings and energy sources.
    '''

    def __init__(self, path, layer = None):
        '''
        Initializes the Streets class with a GeoDataFrame from a specified path.

        Parameters
        ----------
        path : str
            The path to the file containing street geometries.
        layer : str, optional
            The layer to read from the file (default is None).
        '''
        if layer == None:
            self.gdf = gpd.read_file(path)
        else: 
            self.gdf = gpd.read_file(path, layer=layer)

    def add_connection_to_streets(self, buildings, sources):
        '''
        Inserts connection points from buildings and energy sources into the street lines.

        Parameters
        ----------
        buildings : GeoDataFrame
            A GeoDataFrame containing building geometries and attributes, including 'street_id' and 'Anschlusspunkt'.
        sources : GeoDataFrame
            A GeoDataFrame containing energy source geometries and attributes, including 'street_id' and 'Anschlusspunkt'.
        '''

        for df in [buildings, sources]:
            for index, row in df.iterrows():
                street_id = row['street_id']
                if not pd.isna(street_id):
                    anschlusspunkt = row['Anschlusspunkt']
                    line = self.gdf['geometry'][street_id]
                    line_coords = list(line.coords)

                    insertion_position = None
                    min_distance = float('inf')

                    # Find insertion position in the line
                    for i in range(1, len(line_coords)):
                        segment = LineString([line_coords[i-1], line_coords[i]])
                        distance = segment.distance(anschlusspunkt)

                        if distance < min_distance:
                            min_distance = distance
                            insertion_position = i

                    # Insert the connection point into the line coordinates
                    if (anschlusspunkt.x, anschlusspunkt.y) not in line_coords:
                        line_coords.insert(insertion_position, (anschlusspunkt.x, anschlusspunkt.y))
                        self.gdf.at[street_id, 'geometry'] = LineString(line_coords)

class Source:
    '''
    A class to manage energy source geometries and to find the closest points on street networks.

    Attributes
    ----------
    gdf : GeoDataFrame
        A GeoDataFrame containing source geometries and attributes.

    Methods
    -------
    closest_points_sources(streets):
        Finds the closest points on the street network for each energy source and adds these points to the GeoDataFrame.
    '''

    def __init__(self, path, layer = None):
        '''
        Initializes the Source class with a GeoDataFrame from a specified path.

        Parameters
        ----------
        path : str
            The path to the file containing source geometries.
        layer : str, optional
            The layer to read from the file (default is None).
        '''
        if layer == None:
            self.gdf = gpd.read_file(path)
        else: 
            self.gdf = gpd.read_file(path, layer=layer)
        
    def closest_points_sources(self, streets):
        '''
        Finds the closest point on the street network for each energy source and adds these points to the GeoDataFrame.

        Parameters
        ----------
        streets : GeoDataFrame
            A GeoDataFrame containing street geometries and attributes.
        '''
        # Iteration over each source and finding the closest point on the street network
        for index, row_s in self.gdf.iterrows():

            # Initialize variables for minimum distance and closest point
            min_distance = float('inf')
            closest_point = None
            source = row_s['geometry']

            # Iterate over each line in the street network
            for idx,row in streets.iterrows():
                line_coords = list(row['geometry'].coords)  # List of points that make up the line
                
                # Iterate over each line segment to find the closest point
                for i in range(1, len(line_coords)):
                    start_point = Point(line_coords[i-1])
                    end_point = Point(line_coords[i])
                    line_segment = LineString([start_point, end_point])

                    distance = line_segment.distance(source)

                    if distance < min_distance:
                        min_distance = distance
                        closest_point = get_closest_point(LineString([start_point, end_point]), source)
                        id = idx
            self.gdf.at[index, 'Anschlusspunkt'] = closest_point
            self.gdf.at[index, 'street_id'] = int(id)

class Buildings:
    '''
    A class to manage building geometries, add centroids, and find the closest points on street networks.

    Attributes
    ----------
    buildings_all : GeoDataFrame
        A GeoDataFrame containing all building geometries and attributes.
    gdf : GeoDataFrame
        A GeoDataFrame containing buildings with a specified heat attribute greater than zero.

    Methods
    -------
    add_centroid():
        Adds the centroid of each building's geometry to the GeoDataFrame.
    closest_points_buildings(streets):
        Finds the closest point on the street network for each building and adds these points to the GeoDataFrame.
    '''

    def __init__(self, path, heat_att, layer = None):
        '''
        Initializes the Buildings class with a GeoDataFrame from a specified path and filters buildings based on a heat attribute.

        Parameters
        ----------
        path : str
            The path to the file containing building geometries.
        heat_att : str
            The name of the attribute representing heat consumption.
        layer : str, optional
            The layer to read from the file (default is None).
        '''
        if layer == None:
            self.buildings_all = gpd.read_file(path)
        else: 
            self.buildings_all = gpd.read_file(path, layer=layer)
        
        # Filter buildings with heat consumption
        try:
            buildings_wvbr = self.buildings_all[self.buildings_all[heat_att]>0] 
        except:
            print('Check heat attribute!')

        self.gdf = buildings_wvbr
    
    def add_centroid(self):
        '''
        Adds the centroid of each building's geometry to the GeoDataFrame.

        Notes
        -----
        The centroid is computed for each polygon in the GeoDataFrame and added as a new column 'centroid'.
        '''
        self.gdf = self.gdf.copy() # Suppress warning
        self.gdf['centroid'] = self.gdf.loc[:, 'geometry'].centroid

    def closest_points_buildings(self, streets):
        '''
        Finds the closest point on the street network for each building and adds these points to the GeoDataFrame.

        Parameters
        ----------
        streets : GeoDataFrame
            A GeoDataFrame containing street geometries and attributes.

        Notes
        -----
        For each building, this method computes the closest point on the street network and adds it to the GeoDataFrame 
        along with the ID of the closest street.
        '''
        # Create spatial index for the streets
        sindex = streets.sindex

        # Iterate over each building centroid
        for index, row_p in self.gdf.iterrows():
            centroid = row_p['centroid']

            # Use spatial index to get the nearest lines to the centroid
            possible_matches_index = list(sindex.nearest(centroid))
            possible_matches = streets.iloc[[i[0] for i in possible_matches_index]]

            # Find the line closest to the centroid
            closest_line = possible_matches.geometry.distance(centroid).idxmin()

            # Compute the closest point on this line
            closest_point = get_closest_point(streets.at[closest_line, 'geometry'], centroid)

            self.gdf.loc[index, 'Anschlusspunkt'] = closest_point
            self.gdf.loc[index, 'street_id'] = int(closest_line)

class Graph:
    '''
    A class to represent and manipulate a street network graph using NetworkX.

    Attributes
    ----------
    graph : nx.Graph
        A NetworkX graph representing the street network.
    crs : string
        coordinate reference system 

    Methods
    -------
    create_street_network(streets):
        Creates a street network graph from a GeoDataFrame of streets.
    connect_centroids(buildings):
        Connects building centroids to the street network.
    connect_source(sources):
        Connects energy sources to the street network.
    add_attribute_length():
        Adds a 'length' attribute to each edge in the graph.
    plot_G():
        Plots the street network graph.
    get_connected_points(input_point):
        Returns the points connected to the given input point in the graph.
    plot_graph(input_point, connected_points):
        Plots the graph with connected points highlighted.
    graph_to_gdf():
        Converts the NetworkX graph to a GeoDataFrame.
    save_nodes_to_shapefile(filename):
        Saves the graph nodes as points in a shapefile, with node degree and coordinates annotated.
    '''
    def __init__(self, crs):
        '''
        Initializes the Graph class with an empty NetworkX graph.
        '''
        self.graph = nx.Graph()
        self.crs = crs
        
    def create_street_network(self, streets):
        '''
        Creates a street network graph from a GeoDataFrame of streets.

        Parameters
        ----------
        streets : GeoDataFrame
            A GeoDataFrame containing street geometries.
        '''
        # Dictionary with attributes for the edges
        edge_data = {'type': 'Straßenleitung'}  

        # Add nodes and edges
        for idx, row in streets.iterrows():
            geom = row['geometry']
            line_coords = list(geom.coords)

            # Iterate over each point on the line
            for i in range(len(line_coords)):
                node = line_coords[i]
                self.graph.add_node(node)

                # Connect point to previous point
                if i > 0:
                    prev_node = line_coords[i-1]
                    self.graph.add_edge(node, prev_node,**edge_data)
    
    def connect_centroids(self, buildings):
        '''
        Connects building centroids to the street network.

        Parameters
        ----------
        buildings : GeoDataFrame
            A GeoDataFrame containing building geometries and centroids.
        '''
        for index, row in buildings.iterrows():
            centroid = row['centroid']
            closest_point = row['Anschlusspunkt']
            if not pd.isna(closest_point):
                edge_data = {'type': 'Hausanschluss'}  # Dictionary mit dem Attribut, das die edge haben soll
                self.graph.add_edge(centroid.coords[0], (closest_point.x, closest_point.y), **edge_data)

    def connect_source(self, sources):
        '''
        Connects energy sources to the street network.

        Parameters
        ----------
        sources : GeoDataFrame
            A GeoDataFrame containing energy source geometries.
        '''
        for index, row in sources.iterrows():
            source = row['geometry']
            closest_point = row['Anschlusspunkt']
            if not pd.isna(source):
                edge_data = {'type': 'Quellenanschluss'}  # Dictionary mit dem Attribut, das die edge haben soll
                self.graph.add_edge(source.coords[0], (closest_point.x, closest_point.y), **edge_data)

    def add_attribute_length(self):
        '''
        Adds a 'length' attribute to each edge in the graph.
        '''
        for node1, node2 in self.graph.edges():
            geom = LineString([node1, node2])
            self.graph.edges[node1, node2]['length [m]'] = geom.length

    def plot_G(self):
        '''
        Plots the street network graph.
        '''
        # set crs
        pos = {node: (node[0], node[1]) for node in self.graph.nodes}

        plt.figure()
        plt.title('Graph')
        nx.draw_networkx(self.graph, pos=pos, with_labels=False, font_size=6, node_size=3, node_color='blue', edge_color='gray')
        plt.show()

    
    def get_connected_points(self, input_point):
        '''
        Returns the points connected to the given input point in the graph.

        Parameters
        ----------
        input_point : tuple
            The input point coordinates.

        Returns
        -------
        list
            A list of points connected to the input point.
        '''
        # Check input point
        if input_point not in self.graph.nodes:
            print("Input point not in graph nodes.")
            return []

        # Get connected components
        for component in nx.connected_components(self.graph):
            if input_point in component:
                return list(component - {input_point})
        return []

    def plot_graph(self, input_point, connected_points, disconnected_buildings):
        '''
        Plots the graph with connected points highlighted.

        Parameters
        ----------
        input_point : tuple
            The input point coordinates.
        connected_points : list
            A list of points connected to the input point.
        disconnected_buildins : list
            A list of building centroids disconnected from the imput point
        '''
        # Position dictionary for nodes, mapping each node to its (x, y) coordinates for plotting
        pos = {node: (node[0], node[1]) for node in self.graph.nodes}
        # Node colors
        node_colors = [
            '#0000FF' if node == input_point  # blue
            else '#00FF00' if node in connected_points  # green
            else '#800080' if node in disconnected_buildings  # violet
            else '#FFA500' for node in self.graph.nodes]  # orange

        plt.figure(figsize=(20, 20))
        plt.title('Graph Network with connected and disconnected Points')

        # Legend
        legend_labels = {'Source': '#0000FF', 'Connected Points': '#00FF00', 'Disconnected Points': '#FFA500', 'Disconnected Buildings': '#800080'}
        legend_handles = [plt.Line2D([0], [0], marker='o', color=color, label=label, linestyle='None') for label, color in legend_labels.items()]
        plt.legend(handles=legend_handles, loc='upper right', fontsize=10)

        nx.draw(self.graph, pos, node_color=node_colors, font_size=6, node_size=10, with_labels=False)
        plt.show()

        
    def graph_to_gdf(self): # Methode ist ebenfalls in Net. Klassen zusammenfügen? --> Wegen übersichtlichkeit erstmal nicht
        '''
        Converts the NetworkX graph to a GeoDataFrame, including edge attributes.

        Returns
        -------
        GeoDataFrame
            A GeoDataFrame representing the graph edges.
        '''
        geometries = []
        attributes = {}

        for u, v, data in self.graph.edges(data=True):
            geometries.append(LineString([u, v]))

            # Collect attributes for each edge
            for key, value in data.items():
                if key in attributes:
                    attributes[key].append(value)
                else:
                    attributes[key] = [value]

        self.gdf = gpd.GeoDataFrame(attributes, geometry=geometries, crs=self.crs)

    def save_nodes_to_shapefile(self, filename):
        """
        Saves the graph nodes as points in a shapefile, with node degree and coordinates annotated.

        Parameters
        ----------
        filename : str
            The file path to save the shapefile.
        """
        nodes_data = {'geometry': [], 'degree': [], 'x_coord': [], 'y_coord': []}

        for node in self.graph.nodes():
            nodes_data['geometry'].append(Point(node))
            nodes_data['degree'].append(self.graph.degree(node))
            nodes_data['x_coord'].append(node[0])
            nodes_data['y_coord'].append(node[1])

        nodes_gdf = gpd.GeoDataFrame(nodes_data, crs=self.crs)
        nodes_gdf.to_file(filename,driver='GPKG')

class Net:
    '''
    A class to represent and manipulate a network graph for heat distribution.

    Attributes
    ----------
    net : nx.Graph
        A NetworkX graph representing the network.
    htemp : float
        Supply temperature.
    ltemp : float
        Return temperature.
    crs : string
        coordinate reference system

    Methods
    -------
    update_attribute(u, v, attribute, name):
        Adds or updates an attribute to an edge in the network graph.
    add_edge_attributes(pipe_info):
        Adds attributes to the network edges such as GLF, power_th_GLF, volumeflow, DN, velocity, and loss.
    network_analysis(G, buildings, sources, pipe_info, power_th_att, weight='length', progressBar=None):
        Calculates the network by finding the shortest path to each building.
    plot_network(streets, buildings, sources, filename, title='Street network and calculated network'):
        Plots the street network, buildings, and calculated network, and saves the image.
    ensure_power_th_attribute():
        Ensures that each edge in the graph has the thermal power attribute.
    graph_to_gdf():
        Converts a NetworkX graph to a GeoDataFrame, including edge attributes.
    '''

    def __init__(self, htemp, ltemp, crs):
        '''
        Initializes the Net class with an empty NetworkX graph, supply temperature, and return temperature.
        '''
        self.net = nx.Graph()
        self.htemp = htemp
        self.ltemp = ltemp
        self.crs = crs

    def update_attribute(self, u, v, attribute, name):
        '''
        Adds or updates an attribute to an edge in the network graph.

        Parameters
        ----------
        u, v : nodes
            Nodes defining the edge.
        attribute : any
            Value of the attribute.
        name : str
            Name of the attribute.
        '''
        if name in self.net.edges[u, v]:
            self.net.edges[u, v][name] += attribute
        else:
            self.net.edges[u, v][name] = attribute

    def add_edge_attributes(self, pipe_info):
        '''
        Adds attributes to the network edges such as GLF, power_th_GLF, volumeflow, DN, velocity, and loss.

        Parameters
        ----------
        pipe_info : DataFrame
            DataFrame containing pipe information.
        '''
        for (u, v, data) in self.net.edges(data=True):
            n_building = data['n_building']
            power_th = data['power_th [kW]']
            length = data['length [m]']
            edge_type = data.get('type', None)

            GLF = calculate_GLF(n_building)
            power_th_GLF = power_th * GLF
            volumeflow = calculate_volumeflow(power_th_GLF, self.htemp, self.ltemp)
            diameter, velocity, loss, loss_extra = calculate_diameter_velocity_loss(volumeflow, self.htemp, self.ltemp, length, pipe_info, edge_type)
            
            # Add attributes to the edges
            data['GLF'] = GLF
            data['power_th_GLF [kW]'] = power_th_GLF
            data['Volumeflow [l/s]'] = volumeflow
            data['DN [mm]'] = diameter
            data['velocity [m/s]'] = velocity
            data['loss [kWh/a]'] = loss
            data['loss_extra_insulation [kWh/a]'] = loss_extra

    def network_analysis(self, G, buildings, sources, pipe_info, power_th_att, weight='length [m]', progressBar=None):
        '''
        Calculates the network by finding the shortest path to each building.

        Parameters
        ----------
        G : nx.Graph
            The street network graph.
        buildings : GeoDataFrame
            GeoDataFrame of buildings.
        sources : GeoDataFrame
            GeoDataFrame of energy sources.
        pipe_info : DataFrame
            DataFrame containing pipe information.
        power_th_att : str
            Attribute name for thermal power in the buildings GeoDataFrame.
        weight : str, optional
            Edge weight attribute for shortest path calculation (default is 'length [m]').
        progressBar : callable, optional
            Progress bar function (default is None).
        '''

        start_point = (sources['geometry'][0].x, sources['geometry'][0].y)

        for idx, row in buildings.iterrows():
            end_point = (row['centroid'].x, row['centroid'].y)
            power_th = row[power_th_att]
            buildings_count = 1
            try:
                # Shortest path
                path = nx.shortest_path(G, start_point, end_point, weight=weight)
            
                # Add nodes and edges of the path to the network graph
                for i in range(len(path) - 1):
                    u, v = path[i], path[i+1]

                    # Copy all edge attributes
                    self.net.add_edge(u, v, **G.edges[u, v])

                    # Update attributes
                    self.update_attribute(u, v, power_th, 'power_th [kW]')    
                    self.update_attribute(u, v, buildings_count, 'n_building')
            except Exception as e: 
                print(f'No connection for:\n{row}')
                print(f'Error {e}')
                #sys.exit()
            
        # Add GLF, diameter, velocity, and loss attributes
        self.add_edge_attributes(pipe_info)      

    def plot_network(self, streets, buildings, sources, filename, title='Straßennetzwerk und berechnetes Netz'):
        '''
        Plots the street network, buildings, and calculated network, and saves the image.

        Parameters
        ----------
        streets : GeoDataFrame
            GeoDataFrame of streets.
        buildings : GeoDataFrame
            GeoDataFrame of buildings.
        sources : GeoDataFrame
            GeoDataFrame of energy sources.
        filename : str
            File name to save the image.
        title : str, optional
            Title of the plot (default is 'Street network and calculated network').
        '''
        # Node positions
        pos = {node: (node[0], node[1]) for node in self.net.nodes}

        # Create figure and axes
        fig, ax = plt.subplots(figsize=(15, 15))

        # Plot streets
        streets.plot(ax=ax, edgecolor='gray', zorder=1)

        # Plot buildings
        buildings.plot(ax=ax, facecolor='#ff8888', edgecolor='black', zorder=2)

        # Plot energy source as a point
        sources.plot(ax=ax, marker='o', markersize=15, color='green', zorder=3)

        # Plot network
        nx.draw_networkx_edges(self.net, pos=pos, ax=ax, edge_color='blue', width=1.0)

        # Enable grid and axis title
        #ax.grid(True)
        ax.set_title(title)

        # Save plot
        plt.savefig(filename, bbox_inches='tight')

        # Show plot
        plt.show()

    def ensure_power_th_attribute(self):
        """
        Ensures that each edge in the graph has the thermal power attribute.
        If an edge does not have the attribute, it is initialized with a value of 0.
        """
        for u, v in self.net.edges():
            if 'power_th [kW]' not in self.net[u][v]:
                self.net[u][v]['power_th [kW]'] = 0

    def graph_to_gdf(self):
        '''
        Converts a NetworkX graph to a GeoDataFrame, including edge attributes.
        '''
        geometries = []
        attributes = {}

        for u, v, data in self.net.edges(data=True):
            geometries.append(LineString([u, v]))

            # Collect attributes for each edge
            for key, value in data.items():
                if key in attributes:
                    attributes[key].append(value)
                else:
                    attributes[key] = [value]

        # Create a GeoDataFrame from LineString objects and attributes
        self.gdf = gpd.GeoDataFrame(attributes, geometry=geometries, crs=self.crs)

    def rename_columns(self):
        '''
        renames columns of self.gdf.
        '''
        rename_dict = {
            'type': 'Typ',
            'length [m]': 'Laenge [m]',
            'power_th [kW]': 'Leistung_th [kW]',
            'n_building': 'Anzahl Gebaeude',
            'power_th_GLF [kW]': 'Leistung_th_GLF [kW]',
            'Volumeflow [l/s]': 'Volumenstrom [l/s]',
            'velocity [m/s]': 'Geschwindigkeit [m/s]',
            'loss [kWh/a]': 'Verlust [kWh/a]',
            'loss_extra_insulation [kWh/a]': 'Verlust bei extra Daemmung [kWh/a]'
        }
        self.gdf = self.gdf.rename(columns=rename_dict)

class Result:
    '''
    A class to handle and process results for exporting to Excel.

    Attributes
    ----------
    path : str
        Path to the result file.
    data_dict : dict
        Dictionary containing result data.

    Methods
    -------
    create_data_dict(buildings, net, types, dn_list, heat_att, h_temp, l_temp):
        Creates a dictionary for the results to be used in Excel.
    create_df_from_dataDict(net_name='Netz'):
        Converts the dictionary to a result DataFrame.
    save_in_excel(col=0, index_bool=False, sheet_option='replace', sheet='Zusammenfassung'):
        Saves the DataFrame to an Excel sheet.
    '''

    def __init__(self,path):
        '''
        Initializes the Result class with the path to the result file.

        Parameters
        ----------
        path : str
            Path to the result file.
        '''
        self.path = path # path to result file

    def create_data_dict(self, buildings, net, types, dn_list, heat_att, h_temp, l_temp):
        '''
        Creates a dictionary for the results to be used in Excel.

        Parameters
        ----------
        buildings : DataFrame
            DataFrame of buildings.
        net : DataFrame
            DataFrame of the network.
        types : list
            List of building types.
        dn_list : list
            List of possible pipe diameters.
        heat_att : str
            Attribute name for heat demand in the buildings DataFrame.
        h_temp : float
            Supply temperature.
        l_temp : float
            Return temperature.
        '''
        # Helper function
        def summarize_pipes(df, dn_list):
            '''
            Summarizes pipe lengths, losses, and the number of "Hausanschlüsse" per diameter (DN).

            Parameters
            ----------
            df : DataFrame
                Input DataFrame.
            dn_list : list
                List of all possible diameters (DN).
            '''
            # Drop rows where 'DN' is NaN
            df = df.dropna(subset=['DN [mm]'])

            # Initialize an empty DataFrame for the result
            result = pd.DataFrame({'DN [mm]': dn_list}, index=dn_list)
            result['Anzahl Hausanschluesse'] = 0
            result['Hausanschlusslaenge [m]'] = 0
            result['Trassenlaenge [m]'] = 0
            result['Verlust [MWh/a]'] = 0
            result['Verlust bei extra Daemmung [MWh/a]'] = 0

            # Group by DN and type
            grouped = df.groupby(['DN [mm]', 'Typ'])

            # Sum up the pipe lengths and losses, count the number of Hausanschlüsse
            for (dn, pipe_type), group in grouped:
                if pipe_type == 'Hausanschluss':
                    result.loc[dn, 'Hausanschlusslaenge [m]'] += group['Laenge [m]'].sum()
                    result.loc[dn, 'Anzahl Hausanschluesse'] += len(group)
                else:
                    result.loc[dn, 'Trassenlaenge [m]'] += group['Laenge [m]'].sum()

                result.loc[dn, 'Verlust [MWh/a]'] += group['Verlust [MWh/a]'].sum()
                result.loc[dn, 'Verlust bei extra Daemmung [MWh/a]'] += group['Verlust bei extra Daemmung [MWh/a]'].sum()

            return result
        
        # Accumulated building heat demand and count per load profile
        kum_b = buildings.groupby('Lastprofil').agg({heat_att: 'sum'}).reset_index()
        kum_b['count'] = buildings.groupby('Lastprofil').size().reset_index(name='count')['count']

        kum_b[heat_att]/=1000 #MW
        
        # Identify missing load profiles
        missing_types = set(types) - set(kum_b['Lastprofil'])

        if missing_types:
            # Create a DataFrame for missing types
            missing_df = pd.DataFrame({'Lastprofil': list(missing_types), 'count': 0, heat_att: 0})

            # Add missing types
            df = pd.concat([kum_b, missing_df], ignore_index=True)
        else:
            df = kum_b
        
        # Sort
        df_sorted = df.sort_values(by='Lastprofil', key=lambda x: x.map({val: i for i, val in enumerate(types)}))

        # kW in MW
        gdf = net.copy()
        gdf['Leistung_th_GLF [MW]'] = gdf['Leistung_th_GLF [kW]']/1000 
        gdf['Verlust [MWh/a]'] = gdf['Verlust [kWh/a]']/1000 
        gdf['Verlust bei extra Daemmung [MWh/a]'] = gdf['Verlust bei extra Daemmung [kWh/a]']/1000 

        # Length and loss of all diameters in chronological order
        length_loss = summarize_pipes(gdf, dn_list)

        data_dict = {}
        data_dict['Lastprofil'] = types
        data_dict['Anzahl'] = df_sorted['count'].tolist()
        data_dict['Waermebedarf [MWh/a]'] = df_sorted[heat_att].tolist()
        data_dict['DN [mm]'] = dn_list
        data_dict['Anzahl Hausanschluesse'] = length_loss['Anzahl Hausanschluesse'].tolist()
        data_dict['Hausanschlusslaenge [m]'] = length_loss['Hausanschlusslaenge [m]'].tolist()
        data_dict['Trassenlaenge [m]'] = length_loss['Trassenlaenge [m]'].tolist()
        data_dict['Verlust [MWh/a]'] = length_loss['Verlust [MWh/a]'].tolist()
        data_dict['Verlust bei extra Daemmung [MWh/a]'] = length_loss['Verlust bei extra Daemmung [MWh/a]'].tolist()
        data_dict['Vorlauftemp [°C]'] = [h_temp]
        data_dict['Ruecklauftemp [°C]'] = [l_temp]
        data_dict['Max. Leistung (inkl. GLF) [MW]'] = [gdf['Leistung_th_GLF [MW]'].max()]
        data_dict['GLF'] = [calculate_GLF(sum(data_dict['Anzahl']))]

        self.data_dict = data_dict

    def create_df_from_dataDict(self):
        '''
        Converts the dictionary to a result DataFrame.

        Parameters
        ----------
        net_name : str, optional
            Name of the network (default is 'Netz').
        '''
        # Convert the dictionary to a DataFrame
        df = pd.DataFrame.from_dict(self.data_dict, orient='index').transpose()

        # Sum selected columns and write the sum in one row
        sum_row = df[['Anzahl', 'Waermebedarf [MWh/a]', 'Max. Leistung (inkl. GLF) [MW]', 'Hausanschlusslaenge [m]', 'Trassenlaenge [m]', 'Verlust [MWh/a]', 'Verlust bei extra Daemmung [MWh/a]']].sum()
        sum_row['Lastprofil'] = 'Gesamt'
        df_sum = pd.DataFrame([sum_row], columns=['Lastprofil', 'Anzahl', 'Waermebedarf [MWh/a]', 'Max. Leistung (inkl. GLF) [MW]', 'Hausanschlusslaenge [m]', 'Trassenlaenge [m]', 'Verlust [MWh/a]', 'Verlust bei extra Daemmung [MWh/a]'], index=['Summe'])

        # Add the sum row to df
        df = pd.concat([df, df_sum])

        self.df = df
    
    def save_in_excel(self, result_table, col = 0, row = 0, index_bool=False, sheet_option ='replace', sheet = 'Zusammenfassung'):
        '''
        Saves the DataFrame to an Excel sheet.

        Parameters
        ----------
        col : int, optional
            Starting column (default is 0).
        index_bool : bool, optional
            Whether to save the DataFrame with or without indices (default is False).
        sheet_option : str, optional
            Option for handling existing sheets ('replace', 'overlay', or 'new') (default is 'replace').
        sheet : str, optional
            Sheet name in the Excel file (default is 'Zusammenfassung').
        '''
        # Check if the file exists
        if os.path.exists(self.path):
            mode = 'a'  # Append mode
            writer_args = {'if_sheet_exists': sheet_option}
        else:
            mode = 'w'  # Write mode, creates a new file
            writer_args = {}

        # Open the Excel file in the appropriate mode and write the DataFrame to the specified sheet
        with pd.ExcelWriter(self.path, engine='openpyxl', mode=mode, **writer_args) as writer:
            result_table.to_excel(writer, sheet_name=sheet, index=index_bool, startcol=col, startrow=row)
        
        # Adjust column widths
        wb = load_workbook(filename = self.path)        
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length+1)
            ws.column_dimensions[column].width = adjusted_width
        wb.save(self.path)
        
    def is_excel_file_open(self):
        """Checks if the Excel file can be written to."""
        if not os.path.exists(self.path):
            return False  # The file doesn't exist, thus not open
        
        try:
            # Attempt to open the file in read/write mode
            with open(self.path, 'r+'):
                return False  # The file is not open and can be written to
        except IOError:
            # If an IOError occurs, the file may be locked by another application (e.g., Excel)
            return True

    @staticmethod
    def building_statistic(gdf):
        '''
        Computes statistics for buildings from a GeoDataFrame.

        This function filters and aggregates building data to compute various statistics for each building type.
        
        Parameters
        ----------
        gdf : GeoDataFrame
            The GeoDataFrame containing building data. Expected columns include 'type', 'NF', 'RW_spez', 'WW_spez', 'RW_WW_spez', 'age_LANUV', and 'BAK'.

        Returns
        -------
        aggregated_stats : DataFrame
            A DataFrame with aggregated statistics for each building type.
        '''
        filtered_gdf = gdf[gdf['typ'].apply(lambda x: ',' not in x)]
        aggregated_stats = filtered_gdf.groupby('typ').agg(
            NF_median=('NF [m²]', 'median'),                      # Median heated area
            RW_spez_median=('RW_spez [kWh/a*m²]', 'median'),      # Median specific room heat
            WW_spez_median=('WW_spez [kWh/a*m²]', 'median'),      # Median specific warm water
            RW_WW_spez_median=('RW_WW_spez [kWh/a*m²]','median'), # Median combined heating
            Anzahl=('typ', 'size'),                                # Building count
            haeufigstes_Alter_LANUV=('Alter_LANUV', lambda x: x.mode().iloc[0]),  # most common age according to LANUV
            haeufigste_BAK_ALKIS=('BAK nach Flurstueck', lambda x: x.mode().iloc[0]) # most common age (Baualtersklasse) according to ALKIS parcels
        ).reset_index()
        return aggregated_stats
    
    def copy_excel_file(self, source_path):
        """
        Copies an Excel file from source to destination without modifying the content, format, or objects.

        Parameters
        ----------
        source_path : str
            The file path of the source Excel file.
        destination_path : str
            The file path where the Excel file should be saved.
        """
        # Load the workbook from the source file
        workbook = load_workbook(source_path)

        # Save the workbook to the new destination
        workbook.save(self.path)